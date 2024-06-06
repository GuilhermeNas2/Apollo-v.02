import xlwings 
import pandas as pd
import os
import re
from utilsClass import Utils
from dotenv import load_dotenv
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook

class Excel:
     
     global nomeDoArquivo  
     global path   

     load_dotenv()
     path = os.getenv("pathEx")

     dir_list = os.listdir(path)     
    
     nomeDoArquivo = path+dir_list[0]    
     
     def searchExcelBF(data):
        valor_procurado = data
        itensList = []          

        colunaN = "NOTAS" 
        colunaF = "Total Frete"  

        wb = load_workbook(nomeDoArquivo)    

        ctePage = wb['Dados Viagens']
        rowC = None
        
        while len(itensList) <= 1:
            for row in ctePage:            
                for cel in row:
                    if cel.value == colunaF:                                                       
                        itensList.append(cel.column_letter)                        
                    if cel.value == colunaN:
                        itensList.append(cel.column)   
        print(valor_procurado)                

        for row in ctePage.iter_rows(min_col= 4, max_col= 4):
            for cel in row: 
                if type(cel.value) == float:
                    
                    valor_procurado = data+".0"                                         
                if str(cel.value) == str(valor_procurado) and rowC is None:                                    
                    rowC = cel.row 
        if rowC == None:
            return rowC        
          
        notas = Decimal(ctePage.cell(row=rowC, column=itensList[0]).value)
    
        app = xlwings.App(visible=False)  
        wbl = xlwings.Book(nomeDoArquivo)
        
        ctePagel = wbl.sheets['Dados Viagens']
        loc = itensList[1]+str(rowC)        
        frete = Decimal(ctePagel.range(loc).value)        
       
        item = frete/notas
        item = float(item.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        item = "{:.2f}".format(item)  
        
        wb.close()
        app.quit()    
           
        return item  

     def insertExcelN(title,data, valor):
        coll = 1
        maxColl = 26    
        rowl = None  
        cond = True
        count = 1
             
        wb = load_workbook(nomeDoArquivo)      
        ctePage = wb['CTE']    

        padrao = r'\s*(\d+)\s*/' 
        data = re.search(padrao, data)
        title = int(title)

        for row in ctePage.iter_rows(min_col= 1, max_col=26):            
            for cel in row:                
                if cel.value == title:                                                         
                    rowl = cel.row + 1 
                    coll = cel.column + 1
        
        if rowl == None:  
            rowl = 1  
                                                           
            if coll == 1 and ctePage.cell(row=rowl, column=coll).value is not None:
                while count <= 2:   
                    cellTitle = ctePage.cell(row=rowl, column=coll)                
                    if cellTitle.value is None and coll <= maxColl:                    
                        count += 1        
                    elif cellTitle.value is not None and coll <= maxColl:
                        count = 0                
                    if coll <= maxColl:
                        coll += 1    
                    else:    
                        for row in ctePage.iter_rows(min_row=rowl,max_row=rowl,min_col= 1, max_col=26):            
                            for cel in row:
                                if cel.value is not None:                                                                    
                                    coll = 1
                                    rowl +=1
                                    count = 2
                coll -= 1      
            cellTitle = ctePage.cell(row=rowl, column=coll).value = title
            coll += 1           
            cellTitle = ctePage.cell(row=rowl, column=coll).value = "Valor CTE"  
                                
        
        if coll > 1:            
            coll -= 1
        while cond == True:
            cell = ctePage.cell(row=rowl, column=coll)
            
            if cell.value is not None:               
               rowl += 1

            if cell.value is None:                
                cell.value = data.group(1)               
                coll+=1
                cellTwo = ctePage.cell(row=rowl, column=coll)
                if cellTwo.value is None:                   
                    cellTwo.value = "R$"+str(valor)
                    rowl += 1
                    coll = 1
                    cond = False 
                                  
        wb.save(nomeDoArquivo)                      
        return         
                    
                
                    

